# -*- coding: utf-8 -*-
"""团队路由:项目参赛团队 CRUD + Excel 批量导入。

团队信息存 projects/{dir}/teams.json。团队名同时作为 uploads 子目录名(清洗非法字符)。
"""
import io
import uuid
from flask import Blueprint, request, jsonify

from guards import role_required
from storage import read_json, write_json
from config import PROJECTS_DIR
from routes.projects import load_project

bp = Blueprint('teams', __name__, url_prefix='/api')


def _teams_rel(dir_):
    return '{}/{}/teams.json'.format(PROJECTS_DIR, dir_)


def _team_from_data(data):
    return {
        'team_id': 'tm_' + uuid.uuid4().hex[:8],
        'name': (data.get('name') or '').strip(),
        'leader': (data.get('leader') or '').strip(),
        'student_id': (data.get('student_id') or '').strip(),
        'contact': (data.get('contact') or '').strip(),
        'members': (data.get('members') or '').strip(),
        'advisor': (data.get('advisor') or '').strip(),
        'remark': (data.get('remark') or '').strip(),
    }


@bp.get('/projects/<pid>/teams')
@role_required('部长', '副部长')
def list_teams(pid):
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    teams = read_json(_teams_rel(dir_), default=[]) or []
    return jsonify(teams=teams)


@bp.post('/projects/<pid>/teams')
@role_required('部长', '副部长')
def create_team(pid):
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    data = request.get_json(silent=True) or {}
    if not (data.get('name') or '').strip():
        return jsonify(error='团队名称必填'), 400
    teams = read_json(_teams_rel(dir_), default=[]) or []
    team = _team_from_data(data)
    teams.append(team)
    write_json(_teams_rel(dir_), teams)
    return jsonify(team=team)


@bp.put('/projects/<pid>/teams/<tid>')
@role_required('部长', '副部长')
def update_team(pid, tid):
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    data = request.get_json(silent=True) or {}
    teams = read_json(_teams_rel(dir_), default=[]) or []
    team = next((t for t in teams if t.get('team_id') == tid), None)
    if not team:
        return jsonify(error='团队不存在'), 404
    for k in ('name', 'leader', 'student_id', 'contact', 'members', 'advisor', 'remark'):
        if k in data:
            team[k] = (data[k] or '').strip() if isinstance(data[k], str) else data[k]
    write_json(_teams_rel(dir_), teams)
    return jsonify(team=team)


@bp.delete('/projects/<pid>/teams/<tid>')
@role_required('部长', '副部长')
def delete_team(pid, tid):
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    teams = read_json(_teams_rel(dir_), default=[]) or []
    teams = [t for t in teams if t.get('team_id') != tid]
    write_json(_teams_rel(dir_), teams)
    return jsonify(ok=True)


@bp.post('/projects/<pid>/teams/import')
@role_required('部长', '副部长')
def import_teams(pid):
    """批量导入团队。Excel 表头:团队名称 | 队长 | 学号 | 联系方式 | 成员 | 指导老师 | 备注。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    f = request.files.get('file')
    if not f:
        return jsonify(error='未上传文件'), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return jsonify(error='文件解析失败: {}'.format(e)), 400

    teams = read_json(_teams_rel(dir_), default=[]) or []
    existing = {t.get('name') for t in teams}
    added, skipped = 0, 0
    for r in rows:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        if name in existing:
            skipped += 1
            continue
        teams.append({
            'team_id': 'tm_' + uuid.uuid4().hex[:8],
            'name': name,
            'leader': str(r[1]).strip() if len(r) > 1 and r[1] else '',
            'student_id': str(r[2]).strip() if len(r) > 2 and r[2] else '',
            'contact': str(r[3]).strip() if len(r) > 3 and r[3] else '',
            'members': str(r[4]).strip() if len(r) > 4 and r[4] else '',
            'advisor': str(r[5]).strip() if len(r) > 5 and r[5] else '',
            'remark': str(r[6]).strip() if len(r) > 6 and r[6] else '',
        })
        existing.add(name)
        added += 1
    write_json(_teams_rel(dir_), teams)
    return jsonify(added=added, skipped=skipped)
