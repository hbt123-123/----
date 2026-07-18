# -*- coding: utf-8 -*-
"""可编辑汇总工作表:导入 Excel、网页编辑、保存、导出、自动判断列。

工作表存 projects/{dir}/worksheet.json: {columns:[{name,type}], rows:[{列名:值}]}
列类型:text(手动) | auto_file(按团队文件自动判断 是/否)
"""
import io
import os
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from guards import role_required
from storage import read_json, write_json, ensure_dir
from config import PROJECTS_DIR, DATA_DIR
from routes.projects import load_project
from utils.paths import safe_join, sanitize_name

bp = Blueprint('worksheets', __name__, url_prefix='/api')


def _ws_rel(dir_):
    return '{}/{}/worksheet.json'.format(PROJECTS_DIR, dir_)


def _empty():
    return {'columns': [], 'rows': []}


def _team_has_files(dir_, team_name):
    if not team_name:
        return False
    rel = 'projects/{}/uploads/{}'.format(dir_, sanitize_name(team_name))
    abs_dir = safe_join(DATA_DIR, rel)
    if not os.path.isdir(abs_dir):
        return False
    return any(not fn.startswith('.') and os.path.isfile(os.path.join(abs_dir, fn))
               for fn in os.listdir(abs_dir))


def _with_auto_columns(dir_, ws):
    """计算 auto_file 列:按行中团队名列查 uploads 目录有无文件。"""
    columns = ws.get('columns') or []
    rows = ws.get('rows') or []
    auto_cols = [c for c in columns if c.get('type') == 'auto_file']
    if not auto_cols:
        return ws
    team_col = next((c['name'] for c in columns if '团队' in c.get('name', '')), None)
    if not team_col:
        return ws
    rows = [dict(r) for r in rows]
    for row in rows:
        has = _team_has_files(dir_, row.get(team_col, ''))
        for ac in auto_cols:
            row[ac['name']] = '是' if has else '否'
    ws['rows'] = rows
    return ws


@bp.get('/projects/<pid>/worksheet')
@role_required('部长', '副部长', '干事')
def get_worksheet(pid):
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    ws = read_json(_ws_rel(dir_), default=None) or _empty()
    ws = _with_auto_columns(dir_, ws)
    return jsonify(worksheet=ws)


@bp.post('/projects/<pid>/worksheet/import')
@role_required('部长', '副部长')
def import_worksheet(pid):
    """上传 Excel 导入:第一行作列名,后续行作数据。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    f = request.files.get('file')
    if not f:
        return jsonify(error='未上传文件'), 400
    try:
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        return jsonify(error='文件解析失败: {}'.format(e)), 400
    if not rows:
        return jsonify(error='Excel 无数据'), 400

    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    columns = [{'name': h or '列{}'.format(i + 1), 'type': 'text'} for i, h in enumerate(headers)]
    data_rows = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        row = {}
        for i, col in enumerate(columns):
            val = r[i] if i < len(r) else None
            row[col['name']] = '' if val is None else str(val)
        data_rows.append(row)
    ws = {'columns': columns, 'rows': data_rows}
    write_json(_ws_rel(dir_), ws)
    ws = _with_auto_columns(dir_, ws)
    return jsonify(worksheet=ws, imported=len(data_rows))


@bp.put('/projects/<pid>/worksheet')
@role_required('部长', '副部长')
def save_worksheet(pid):
    """保存工作表(列定义 + 行数据)。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    data = request.get_json(silent=True) or {}
    columns = data.get('columns') or []
    rows = data.get('rows') or []
    for c in columns:
        if not isinstance(c, dict):
            continue
        c['name'] = str(c.get('name') or '').strip() or '未命名'
        c['type'] = c['type'] if c.get('type') in ('text', 'auto_file') else 'text'
    ws = {'columns': columns, 'rows': rows}
    write_json(_ws_rel(dir_), ws)
    ws = _with_auto_columns(dir_, ws)
    return jsonify(worksheet=ws)


@bp.get('/projects/<pid>/worksheet/download')
@role_required('部长', '副部长')
def download_worksheet(pid):
    """导出 xlsx(含 auto_file 计算结果)。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    ws = read_json(_ws_rel(dir_), default=None) or _empty()
    ws = _with_auto_columns(dir_, ws)

    wb = Workbook()
    sheet = wb.active
    sheet.title = '汇总表'
    headers = [c['name'] for c in ws.get('columns', [])]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in ws.get('rows', []):
        sheet.append([row.get(h, '') for h in headers])

    export_rel = '{}/{}/exports/工作表.xlsx'.format(PROJECTS_DIR, dir_)
    ensure_dir('{}/{}/exports'.format(PROJECTS_DIR, dir_))
    abs_path = safe_join(DATA_DIR, export_rel)
    wb.save(abs_path)
    return send_file(abs_path, as_attachment=True,
                     download_name='{}-工作表.xlsx'.format(sanitize_name(meta.get('name', ''))))
