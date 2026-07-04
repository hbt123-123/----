# -*- coding: utf-8 -*-
"""成员管理路由:CRUD、批量导入、密码重置(仅部长)。"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify

from auth import (
    load_users, save_users, find_user_in, to_public, new_user_id,
)
from guards import admin_required

bp = Blueprint('users', __name__, url_prefix='/api/users')


@bp.get('')
@admin_required
def list_users():
    return jsonify(users=[to_public(u) for u in load_users()])


@bp.post('')
@admin_required
def create_user():
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    role = data.get('role')
    if not name:
        return jsonify(error='姓名必填'), 400
    if role not in ('部长', '副部长', '干事'):
        return jsonify(error='角色非法'), 400
    users = load_users()
    u = {
        'id': new_user_id(),
        'name': name,
        'role': role,
        'student_id': (data.get('student_id') or '').strip(),
        'contact': (data.get('contact') or '').strip(),
        'is_admin': False,
        'activated': False,
        'password_hash': '',
        'created_at': datetime.now().isoformat(),
    }
    users.append(u)
    save_users(users)
    return jsonify(user=to_public(u))


@bp.put('/<uid>')
@admin_required
def update_user(uid):
    data = request.get_json(silent=True) or {}
    users = load_users()
    u = find_user_in(users, uid)
    if not u:
        return jsonify(error='用户不存在'), 404
    if u.get('is_admin') and data.get('role') and data['role'] != '部长':
        return jsonify(error='不可更改管理员角色'), 400
    for k in ('name', 'role', 'student_id', 'contact'):
        if k in data:
            u[k] = data[k]
    save_users(users)
    return jsonify(user=to_public(u))


@bp.delete('/<uid>')
@admin_required
def delete_user(uid):
    users = load_users()
    u = find_user_in(users, uid)
    if not u:
        return jsonify(error='用户不存在'), 404
    if u.get('is_admin'):
        return jsonify(error='不可删除管理员账号'), 400
    users = [x for x in users if x.get('id') != uid]
    save_users(users)
    return jsonify(ok=True)


@bp.post('/<uid>/reset-password')
@admin_required
def reset_password(uid):
    """部长重置成员密码:清除哈希并取消激活,成员下次登录重新设置密码。"""
    users = load_users()
    u = find_user_in(users, uid)
    if not u:
        return jsonify(error='用户不存在'), 404
    if u.get('is_admin'):
        return jsonify(error='不可重置管理员密码'), 400
    u['password_hash'] = ''
    u['activated'] = False
    save_users(users)
    return jsonify(ok=True)


@bp.post('/import')
@admin_required
def import_users():
    """批量导入成员。Excel 表头:姓名 | 学号 | 角色 | 联系方式,从第二行起。"""
    f = request.files.get('file')
    if not f:
        return jsonify(error='未上传文件'), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return jsonify(error='文件解析失败: {}'.format(e)), 400

    users = load_users()
    existing = {(u.get('name'), u.get('student_id')) for u in users}
    added, skipped = 0, 0
    for r in rows:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        sid = str(r[1]).strip() if len(r) > 1 and r[1] else ''
        role = str(r[2]).strip() if len(r) > 2 and r[2] else '干事'
        contact = str(r[3]).strip() if len(r) > 3 and r[3] else ''
        if role not in ('部长', '副部长', '干事'):
            role = '干事'
        if (name, sid) in existing:
            skipped += 1
            continue
        users.append({
            'id': new_user_id(),
            'name': name,
            'role': role,
            'student_id': sid,
            'contact': contact,
            'is_admin': False,
            'activated': False,
            'password_hash': '',
            'created_at': datetime.now().isoformat(),
        })
        existing.add((name, sid))
        added += 1
    save_users(users)
    return jsonify(added=added, skipped=skipped)
