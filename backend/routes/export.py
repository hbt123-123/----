# -*- coding: utf-8 -*-
"""汇总导出:预览汇总数据 + 下载 Excel。

汇总表:行=团队,列= 团队信息 + 各材料状态 + 备注。
生成 xlsx 存 projects/{dir}/exports/,同时返回文件下载。
"""
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from guards import role_required
from storage import read_json, ensure_dir
from config import PROJECTS_DIR, DATA_DIR
from routes.projects import load_project
from utils.paths import safe_join, sanitize_name

bp = Blueprint('export', __name__, url_prefix='/api')


def _build_summary(dir_):
    """构建汇总:materials 列定义 + rows(每团队一行)。

    答辩场次的成绩作为附加列并入汇总表(列名:答辩成绩-<阶段名>)。
    """
    teams = read_json('{}/{}/teams.json'.format(PROJECTS_DIR, dir_), default=[]) or []
    tasks = read_json('{}/{}/tasks.json'.format(PROJECTS_DIR, dir_), default=[]) or []
    defenses = read_json('{}/{}/defenses.json'.format(PROJECTS_DIR, dir_), default=None) or {}
    defense_sessions = defenses.get('sessions', []) if isinstance(defenses, dict) else []

    # 收集材料列(去重,按出现顺序)
    materials = []
    seen = set()
    for t in tasks:
        key = (t.get('stage_id'), t.get('material'))
        if key not in seen:
            seen.add(key)
            materials.append({
                'stage_id': t.get('stage_id', ''),
                'stage_name': t.get('stage_name', ''),
                'material': t.get('material', ''),
            })

    # 答辩场次列(去重,按场次顺序)
    defense_cols = []
    for s in defense_sessions:
        if not s.get('stage_name'):
            continue
        col_name = '答辩成绩-{}'.format(s.get('stage_name', ''))
        if any(c['name'] == col_name for c in defense_cols):
            continue
        defense_cols.append({
            'name': col_name,
            'session_id': s.get('session_id', ''),
            'stage_name': s.get('stage_name', ''),
        })

    rows = []
    for i, team in enumerate(teams):
        statuses = []
        remarks = []
        for m in materials:
            task = next((t for t in tasks if t.get('team_id') == team.get('team_id')
                         and t.get('stage_id') == m['stage_id']
                         and t.get('material') == m['material']), None)
            status = task.get('status', '未交') if task else '未交'
            statuses.append(status)
            if task and task.get('review_comment'):
                remarks.append('{}: {}'.format(m['material'], task['review_comment']))
        # 答辩成绩
        defense_scores = []
        for dc in defense_cols:
            session = next((s for s in defense_sessions if s.get('session_id') == dc['session_id']), None)
            arr = None
            if session:
                arr = next((a for a in session.get('arrangements', [])
                            if a.get('team_id') == team.get('team_id')), None)
            score = arr.get('score', 0) if arr else 0
            defense_scores.append(score if score else '')
        rows.append({
            'no': i + 1,
            'name': team.get('name', ''),
            'leader': team.get('leader', ''),
            'student_id': team.get('student_id', ''),
            'contact': team.get('contact', ''),
            'members': team.get('members', ''),
            'advisor': team.get('advisor', ''),
            'statuses': statuses,
            'defense_scores': defense_scores,
            'remark': '; '.join(remarks),
        })
    return materials, rows, defense_cols


@bp.get('/projects/<pid>/export/preview')
@role_required('部长', '副部长')
def preview_summary(pid):
    """汇总预览(供前端表格展示)。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    materials, rows, defense_cols = _build_summary(dir_)
    return jsonify(materials=materials, rows=rows, defense_cols=defense_cols, project_name=meta.get('name', ''))


@bp.get('/projects/<pid>/export/download')
@role_required('部长', '副部长')
def download_summary(pid):
    """生成并下载 Excel 汇总表,同时存项目 exports/ 目录。"""
    meta, dir_ = load_project(pid, request.current_user)
    if not dir_:
        return jsonify(error='项目不存在或无权访问'), 404
    materials, rows, defense_cols = _build_summary(dir_)

    wb = Workbook()
    ws = wb.active
    ws.title = '汇总表'

    headers = ['序号', '团队名称', '队长', '学号', '联系方式', '成员', '指导老师']
    for m in materials:
        headers.append('{}-{}'.format(m['stage_name'], m['material']))
    for dc in defense_cols:
        headers.append(dc['name'])
    headers.append('备注')
    ws.append(headers)

    # 表头样式
    head_fill = PatternFill('solid', fgColor='ECF5FF')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in rows:
        ws.append([row['no'], row['name'], row['leader'], row['student_id'],
                   row['contact'], row['members'], row['advisor']]
                  + row['statuses'] + row['defense_scores'] + [row['remark']])

    # 列宽
    widths = [6, 16, 10, 12, 13, 22, 10] + [14] * len(materials) + [14] * len(defense_cols) + [26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 存 exports/
    export_rel = '{}/{}/exports/汇总表.xlsx'.format(PROJECTS_DIR, dir_)
    ensure_dir('{}/{}/exports'.format(PROJECTS_DIR, dir_))
    abs_path = safe_join(DATA_DIR, export_rel)
    wb.save(abs_path)

    fname = '{}-汇总表.xlsx'.format(sanitize_name(meta.get('name', '项目')))
    return send_file(abs_path, as_attachment=True, download_name=fname)
